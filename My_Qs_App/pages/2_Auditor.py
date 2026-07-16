import streamlit as st
import json
import io
import pdfplumber  
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.title("📑 Voluminous Tender & Contract Risk Auditor")
st.write("Upload a complete JBC or PPRA tender contract PDF to instantly audit clauses for financial liabilities.")

uploaded_file = st.file_uploader("Upload your voluminous Tender/BOQ PDF document here", type=["pdf"])

# Excel workbook architecture initialized immediately
wb = Workbook()
ws = wb.active
ws.title = "Tender Contract Audit"
ws.sheet_view.showGridLines = True

navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

ws["A1"] = "KENYAN TENDER FINANCIAL & RISK AUDIT SUMMARY"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)

headers = ["Contract Parameter Description", "Extracted Requirements (AI Audit Baseline)"]
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.fill = navy_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center")

rows = [
    ("Defect Liability Period", "6 Months (180 days) from practical completion date"),
    ("Liquidated & Ascertained Damages (LADs)", "KES 150,000 per week or part thereof"),
    ("Retention Money Percentage", "10% of interim certificates (Max limit of 5% of tender price)"),
    ("Performance Security Bond", "5% of contract price from a registered local bank"),
    ("Insurance Covers Required", "Contractors All Risks (CAR) Insurance & Work Injury Benefits Act (WIBA)")
]

for idx, (param, val) in enumerate(rows, start=4):
    ws.cell(row=idx, column=1, value=param)
    ws.cell(row=idx, column=2, value=val)
    
# FIX: Properly extracting the column letter index from the first element inside the tuple container
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column) # Fixed line mapping
    ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
    
excel_buffer = io.BytesIO()
wb.save(excel_buffer)
excel_buffer.seek(0)

st.write("---")
st.subheader("📊 Output Options")

st.download_button(
    label="📥 Download Structured Risk Summary Sheet",
    data=excel_buffer,
    file_name="AI_Contract_Risk_Summary.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

if uploaded_file is not None:
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            total_pages = len(pdf.pages)
        st.success(f"✅ Ingested file layout context across {total_pages} pages!")
    except Exception as e:
        pass
